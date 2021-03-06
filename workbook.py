from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import stock_info as si
from datetime import datetime, timedelta


# creates a workbook with the given name and fills column names
def create_workbook(workbook):
    wb = Workbook()
    ws = wb.active

    # enters in column labels and styles
    ft = Font(bold=True)
    ws.column_dimensions['A'].width = 10.25
    ws['A2'] = 'DATE'
    ws['A2'].font = ft
    ws['B1'] = 'STOCKS'
    ws['B1'].font = ft
    ws['C1'] = 'T. DIF'
    ws['C1'].font = ft
    ws['D1'] = 'T. VALUE'
    ws['D1'].font = ft

    wb.save(workbook)

# grabs the names and quantities of stocks stored in the given worksheet and returns it as a dictionary
def get_stocks(workbook):
    wb = load_workbook(workbook, read_only=True)
    ws = wb.active

    stocks = {}
    col = 66
    while ws[chr(col) + '2'].value is not None:
        stocks[ws[chr(col) + '2'].value.split()[0]] = float(ws[chr(col) + '2'].value.split()[1])
        col += 1
    return stocks


# adds a column into the worksheet with the given quantity
def add_stock(workbook, stock, quantity):
    wb = load_workbook(workbook)
    ws = wb.active

    # checks to see if the stock is already in the sheet
    col = 66
    while ws[chr(col) + '2'].value is not None and ws[chr(col) + '2'].value.split()[0] != stock:
        col += 1
    if ws[chr(col) + '2'].value is not None:
        return 0

    # if this is the first stock to be added to the sheet, it doesn't create a new column
    if col != 66:
        ws.insert_cols(2)
    ws['B2'] = stock + ' ' + str(quantity)

    # moves the 'STOCKS' label over to the correct cell
    ws['B1'] = 'STOCKS'
    ws['B1'].font = Font(bold=True)
    if ws['C1'].value == 'STOCKS':
        ws['C1'] = ''

    wb.save(workbook)
    return 1


# deletes the given stock from the worksheet
def del_stock(workbook, stock):
    wb = load_workbook(workbook)
    ws = wb.active

    # finds the stock or returns 0 if the stock doesn't exist in the sheet
    col = 66
    while ws[chr(col) + '2'].value is not None and ws[chr(col) + '2'].value.split()[0] != stock:
        col += 1
    if ws[chr(col) + '2'].value is None:
        return 0

    # deletes the column and makes sure 'STOCKS' is still in the right place
    ws.delete_cols(col - 64)
    if ws['B1'].value == 'T. DIF':
        ws.insert_cols(2)
    ws['B1'] = 'STOCKS'
    ws['B1'].font = Font(bold=True)

    wb.save(workbook)
    return 1


# changes the quantity of the given stock from the worksheet
def change_stock_q(workbook, stock, quantity):
    wb = load_workbook(workbook)
    ws = wb.active

    # finds the stock or returns 0 if the stock doesn't exist in the sheet
    col = 66
    while ws[chr(col) + '2'].value is not None and ws[chr(col) + '2'].value.split()[0] != stock:
        col += 1
    if ws[chr(col) + '2'].value is None:
        return 0

    # resets the quantity to the given one
    ws[chr(col) + '2'] = stock + ' ' + str(quantity)

    wb.save(workbook)
    return 1


# adds stock info for the given date and stocks to the table
def add_date(workbook, stocks, stocks_q, date):
    wb = load_workbook(workbook)
    ws = wb.active

    # find next open cell in col 'A'
    row = 3
    while ws['A' + str(row)].value is not None:
        row += 1

    row = str(row)

    # enters the date and the stock info into the table and stores total in final column
    gain_loss = 0
    ws['A' + row] = date
    ws['A' + row].number_format = 'm/d/yyyy'
    i = 0
    for i in range(len(stocks)):
        col = chr(66 + i)
        change = si.get_change(stocks[i], date)
        ws[col + row] = change
        ws[col + row].number_format = '[Color10]"$"#,##0.00;[Red]"$"#,##0.00'
        gain_loss += change * stocks_q[stocks[i]]
    ws[chr(67 + i) + row] = gain_loss
    ws[chr(67 + i) + row].number_format = '[Color10]"$"#,##0.00;[Red]"$"#,##0.00'

    # adds the total value of all stocks into the sheet
    total = 0
    for s in stocks:
        total += si.get_close(s, date) * stocks_q[s]
    ws[chr(68 + i) + row] = total
    ws[chr(68 + i) + row].number_format = '"$"#,##0.00'

    wb.save(workbook)
    return 1


# fills in missing dates between the most recent entry and the day before runtime
def fill_dates(workbook):
    wb = load_workbook(workbook)
    ws = wb.active

    stocks_q = get_stocks(workbook)
    stocks = list(stocks_q.keys())

    # gets the number of missing days between the last date and yesterday
    row = 3
    while ws['A' + str(row)].value is not None:
        row += 1
    last_date = ws['A' + str(row - 1)].value
    days = ((datetime.now() - timedelta(1)) - last_date).days

    # calls add_date for missing dates while omitting weekends and days market isn't open (holidays)
    for day in (last_date + timedelta(n + 1) for n in range(days)):
        try:
            if day.weekday() != 5 and day.weekday() != 6:
                add_date(workbook, stocks, stocks_q, day)
        except IndexError:
            pass
