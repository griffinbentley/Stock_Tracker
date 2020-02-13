import window as win
import workbook as wb
from openpyxl import load_workbook
import os


# sets the name of the workbook and opens the workbook editing window
workbook_title = 'Stocks.xlsx'

# runs window allowing for the user to add delete and change stocks and quantities
win.main_window(workbook_title)

# checks to see if there are any stocks in the workbook
workbook = load_workbook(workbook_title, read_only=True)
ws = workbook.active
if ws['B2'].value is not None:

    # checks to see if a start date is needed
    if ws['A3'].value is None:
        stocks_q = wb.get_stocks(workbook_title)
        stocks = list(stocks_q.keys())
        win.init_date_win(workbook_title, stocks, stocks_q)

    # checks to see if the user entered a start date or just closed the window
    workbook = load_workbook(workbook_title, read_only=True)
    ws = workbook.active

    if ws['A3'].value is not None:

        # fills out any missing dates form the last date entered to the day before runtime
        wb.fill_dates(workbook_title)

        # opens the workbook with all updates
        os.startfile(workbook_title)
